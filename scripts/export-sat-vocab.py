# -*- coding: utf-8 -*-
import json
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl

path = Path(
    r"c:\Users\emrea\Downloads\SAT_991_20_Words_Detailed_Study_Plan_TR (1).xlsx"
)
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["Temaya Göre"]
words = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[1]:
        continue
    words.append(
        {
            "no": int(row[0]) if row[0] else len(words) + 1,
            "word": str(row[1]).strip(),
            "pos": str(row[2] or "").strip(),
            "definition": str(row[3] or "").strip(),
            "theme": str(row[4] or "").strip(),
            "study_split": str(row[5] or "").strip(),
            "prefix": str(row[6] or "").strip(),
            "core_stem": str(row[7] or "").strip(),
            "root_family": str(row[8] or "").strip(),
            "root_meaning": str(row[9] or "").strip(),
            "suffix": str(row[10] or "").strip(),
            "morphology_note": str(row[11] or "").strip(),
            "turkish": str(row[12] or "").strip(),
            "detailed_definition_en": str(row[13] or "").strip(),
            "detailed_definition_tr": str(row[14] or "").strip()
            if len(row) > 14 and row[14]
            else "",
            "example_pattern": str(row[15] or "").strip()
            if len(row) > 15 and row[15]
            else "",
        }
    )

ws3 = wb["10 Haftalık Plan"]
plan_rows = []
for row in ws3.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    week = int(row[0])
    day = str(row[1] or "").strip()
    session = str(row[2] or "").strip()
    theme = str(row[3] or "").strip()
    count = row[4]
    words_cell = str(row[5] or "").strip()
    session_num = None
    m = re.search(r"Oturum\s+(\d+)", session, re.I)
    if m:
        kind = "learn"
        session_num = int(m.group(1))
        raw_parts = [x.strip() for x in words_cell.split(",") if x.strip()]
        word_list = []
        extra_notes = []
        for part in raw_parts:
            # Some cells append "word | Ek görev: ..."
            if "|" in part:
                word_part, note_part = part.split("|", 1)
                word_part = word_part.strip()
                if word_part:
                    word_list.append(word_part)
                if note_part.strip():
                    extra_notes.append(note_part.strip())
            else:
                word_list.append(part)
        task_note = " · ".join(extra_notes)
    elif "Dinlenme" in session:
        kind = "rest"
        word_list = []
        task_note = words_cell
    else:
        kind = "review"
        word_list = []
        task_note = words_cell
    plan_rows.append(
        {
            "week": week,
            "day_name": day,
            "session_label": session,
            "session_num": session_num,
            "kind": kind,
            "theme_focus": theme,
            "word_count": len(word_list)
            if kind == "learn"
            else (int(count) if isinstance(count, (int, float)) else None),
            "words": word_list,
            "task_note": task_note,
        }
    )

start = date(2026, 7, 31)
for i, p in enumerate(plan_rows):
    p["scheduled_date"] = (start + timedelta(days=i)).isoformat()
    p["id"] = f"plan-{i + 1:03d}"

themes = []
seen: dict[str, dict] = {}
for w in words:
    t = w["theme"]
    if t not in seen:
        seen[t] = {"theme": t, "order": len(themes) + 1, "word_count": 0}
        themes.append(seen[t])
    seen[t]["word_count"] += 1

out = {
    "meta": {
        "title": "SAT Vocabulary — 10-week plan (20 words/day)",
        "source": "SAT_991_20_Words_Detailed_Study_Plan_TR",
        "word_count": len(words),
        "theme_count": len(themes),
        "plan_start": start.isoformat(),
        "learn_sessions": sum(1 for p in plan_rows if p["kind"] == "learn"),
        "review_days": sum(1 for p in plan_rows if p["kind"] == "review"),
        "rest_days": sum(1 for p in plan_rows if p["kind"] == "rest"),
    },
    "themes": themes,
    "words": words,
    "plan": plan_rows,
}
out_path = Path(
    r"c:\Users\emrea\Cursor\optimizationnextjs\src\lib\sat-vocab\data.json"
)
out_path.parent.mkdir(parents=True, exist_ok=True)
out_path.write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")
catalog_path = out_path.parent / "catalog.json"
words_path = out_path.parent / "words.json"
catalog_path.write_text(
    json.dumps({"meta": out["meta"], "themes": out["themes"], "plan": out["plan"]}, ensure_ascii=False),
    encoding="utf-8",
)
words_path.write_text(json.dumps({"words": out["words"]}, ensure_ascii=False), encoding="utf-8")
print(out["meta"])
print(
    "kinds",
    {k: sum(1 for p in plan_rows if p["kind"] == k) for k in ("learn", "review", "rest")},
)
print("size_mb", round(out_path.stat().st_size / 1e6, 2))
wset = {x["word"].lower() for x in words}
missing = []
for p in plan_rows:
    if p["kind"] != "learn":
        continue
    for ww in p["words"]:
        if ww.lower() not in wset:
            missing.append(ww)
print("missing", len(missing), missing[:15])
